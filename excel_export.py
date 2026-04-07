"""
Excel Export — Generates results spreadsheet.

Format:
  Row 1: Headers (Student Number, Q1 Score, Q1 Confidence, Q2 Score, ...)
  Row 2+: One row per student

Also includes a summary sheet with class statistics.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Any
import os


# Style constants
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
WARNING_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
ERROR_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_results(
    exam_id: str,
    question_nums: List[str],
    student_results: List[Dict[str, Any]],
    output_dir: str = "output",
) -> str:
    """
    Export evaluation results to Excel.

    student_results: list of {
        "studentNumber": "1234567890",
        "studentNumberConfidence": 0.95,
        "questions": {
            "1": {"score": 10, "maxPoints": 10, "confidence": 0.98, "status": "correct", "needsReview": False},
            "2": {"score": 0, "maxPoints": 20, "confidence": 0.45, "status": "pending_ai", "needsReview": True},
            ...
        },
        "totalScore": 70,
        "totalMaxPoints": 100,
    }
    """
    wb = Workbook()

    # ── Sheet 1: Scores ──
    ws = wb.active
    ws.title = "Scores"

    # Header row
    headers = ["#", "Student Number", "SN Confidence"]
    for qn in question_nums:
        headers.extend([f"Q{qn} Score", f"Q{qn} Conf", f"Q{qn} Detail"])
    headers.extend(["Total Score", "Max Points", "Percentage", "Needs Review"])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, student in enumerate(student_results, 2):
        col = 1

        # Row number
        ws.cell(row=row_idx, column=col, value=row_idx - 1).border = THIN_BORDER
        col += 1

        # Student number
        sn_cell = ws.cell(row=row_idx, column=col, value=student.get("studentNumber", "?"))
        sn_cell.border = THIN_BORDER
        col += 1

        # SN confidence
        sn_conf = student.get("studentNumberConfidence", 0)
        sn_conf_cell = ws.cell(row=row_idx, column=col, value=round(sn_conf * 100, 1))
        sn_conf_cell.border = THIN_BORDER
        if sn_conf < 0.80:
            sn_conf_cell.fill = WARNING_FILL
        col += 1

        # Per-question scores
        questions = student.get("questions", {})
        needs_review = False

        for qn in question_nums:
            q_result = questions.get(qn, {})
            score = q_result.get("score", 0)
            conf = q_result.get("confidence", 0)
            explanation = q_result.get("explanation", q_result.get("status", ""))
            status = q_result.get("status", "unknown")
            review = q_result.get("needsReview", False)

            if review or conf < 0.80:
                needs_review = True

            # Score
            score_cell = ws.cell(row=row_idx, column=col, value=score)
            score_cell.border = THIN_BORDER
            if status == "correct":
                score_cell.fill = SUCCESS_FILL
            elif status in ("wrong", "partial"):
                score_cell.fill = ERROR_FILL
            elif status in ("pending_ai", "blank"):
                score_cell.fill = WARNING_FILL
            col += 1

            # Confidence
            conf_cell = ws.cell(row=row_idx, column=col, value=round(conf * 100, 1))
            conf_cell.border = THIN_BORDER
            if conf < 0.80:
                conf_cell.fill = WARNING_FILL
            col += 1

            # Detail (explanation)
            detail_cell = ws.cell(row=row_idx, column=col, value=explanation)
            detail_cell.border = THIN_BORDER
            col += 1

        # Totals
        total = student.get("totalScore", 0)
        max_pts = student.get("totalMaxPoints", 0)
        pct = round(total / max_pts * 100, 1) if max_pts > 0 else 0

        ws.cell(row=row_idx, column=col, value=round(total, 2)).border = THIN_BORDER
        col += 1
        ws.cell(row=row_idx, column=col, value=max_pts).border = THIN_BORDER
        col += 1
        ws.cell(row=row_idx, column=col, value=pct).border = THIN_BORDER
        col += 1

        review_cell = ws.cell(row=row_idx, column=col, value="YES" if needs_review else "")
        review_cell.border = THIN_BORDER
        if needs_review:
            review_cell.fill = WARNING_FILL
            review_cell.font = Font(bold=True, color="856404")

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 20)

    # ── Sheet 2: Summary ──
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Exam ID").font = HEADER_FONT
    ws2.cell(row=1, column=2, value=exam_id)
    ws2.cell(row=2, column=1, value="Total Students").font = HEADER_FONT
    ws2.cell(row=2, column=2, value=len(student_results))
    ws2.cell(row=3, column=1, value="Questions").font = HEADER_FONT
    ws2.cell(row=3, column=2, value=len(question_nums))

    if student_results:
        scores = [s.get("totalScore", 0) for s in student_results]
        max_pts = student_results[0].get("totalMaxPoints", 0)
        ws2.cell(row=5, column=1, value="Average Score").font = HEADER_FONT
        ws2.cell(row=5, column=2, value=round(sum(scores) / len(scores), 2))
        ws2.cell(row=6, column=1, value="Highest Score").font = HEADER_FONT
        ws2.cell(row=6, column=2, value=round(max(scores), 2))
        ws2.cell(row=7, column=1, value="Lowest Score").font = HEADER_FONT
        ws2.cell(row=7, column=2, value=round(min(scores), 2))
        ws2.cell(row=8, column=1, value="Max Points").font = HEADER_FONT
        ws2.cell(row=8, column=2, value=max_pts)

        reviews = sum(1 for s in student_results if any(
            q.get("needsReview") or q.get("confidence", 1) < 0.80
            for q in s.get("questions", {}).values()
        ))
        ws2.cell(row=10, column=1, value="Papers Needing Review").font = HEADER_FONT
        ws2.cell(row=10, column=2, value=reviews)

    # Save
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"{exam_id}_results.xlsx")
    wb.save(filepath)
    print(f"[Excel] Saved to {filepath}")
    return filepath
