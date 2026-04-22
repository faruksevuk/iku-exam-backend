"""
Excel export — per-student rows + summary sheet.

Columns:
  # | Student Number | SN Conf | Q1 Score | Q1 Conf | Q1 Detail | ... | Total | Max | % | Needs Review
"""

import os
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config


HEADER_FONT = Font(bold=True, size=11)
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
ERROR_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def export_results(
    exam_id: str,
    question_nums: List[str],
    student_results: List[Dict[str, Any]],
    output_dir: str = None,
) -> str:
    """Write results to an xlsx file and return its absolute path."""
    output_dir = output_dir or config.OUTPUT_DIR
    wb = Workbook()

    # ── Scores sheet ──
    ws = wb.active
    ws.title = "Scores"

    headers = ["#", "Student Number", "SN Confidence"]
    for qn in question_nums:
        headers.extend([f"Q{qn} Score", f"Q{qn} Conf", f"Q{qn} Detail"])
    headers.extend(["Total Score", "Max Points", "Percentage", "Needs Review"])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for row_idx, student in enumerate(student_results, 2):
        col = 1

        ws.cell(row=row_idx, column=col, value=row_idx - 1).border = THIN_BORDER
        col += 1

        ws.cell(row=row_idx, column=col, value=student.get("studentNumber", "?")).border = THIN_BORDER
        col += 1

        sn_conf = student.get("studentNumberConfidence", 0.0)
        sn_cell = ws.cell(row=row_idx, column=col, value=round(sn_conf * 100, 1))
        sn_cell.border = THIN_BORDER
        if sn_conf < config.HIGH_CONF_THRESHOLD:
            sn_cell.fill = WARNING_FILL
        col += 1

        questions = student.get("questions", {})
        needs_review = False

        for qn in question_nums:
            q = questions.get(qn, {})
            score = q.get("score", 0)
            conf = q.get("confidence", 0.0)
            detail = q.get("explanation", q.get("status", ""))
            status = q.get("status", "unknown")
            review = q.get("needsReview", False)

            if review or conf < config.HIGH_CONF_THRESHOLD:
                needs_review = True

            score_cell = ws.cell(row=row_idx, column=col, value=score)
            score_cell.border = THIN_BORDER
            if status == "correct":
                score_cell.fill = SUCCESS_FILL
            elif status in ("wrong", "partial"):
                score_cell.fill = ERROR_FILL
            elif status in ("pending_review", "blank", "ai_flagged", "pending_ai"):
                score_cell.fill = WARNING_FILL
            col += 1

            conf_cell = ws.cell(row=row_idx, column=col, value=round(conf * 100, 1))
            conf_cell.border = THIN_BORDER
            if conf < config.HIGH_CONF_THRESHOLD:
                conf_cell.fill = WARNING_FILL
            col += 1

            ws.cell(row=row_idx, column=col, value=detail).border = THIN_BORDER
            col += 1

        total = student.get("totalScore", 0)
        max_pts = student.get("totalMaxPoints", 0)
        pct = round(total / max_pts * 100, 1) if max_pts > 0 else 0

        ws.cell(row=row_idx, column=col, value=round(total, 2)).border = THIN_BORDER
        col += 1
        ws.cell(row=row_idx, column=col, value=max_pts).border = THIN_BORDER
        col += 1
        ws.cell(row=row_idx, column=col, value=pct).border = THIN_BORDER
        col += 1

        review_cell = ws.cell(row=row_idx, column=col, value="YES" if needs_review else "")
        review_cell.border = THIN_BORDER
        if needs_review:
            review_cell.fill = WARNING_FILL
            review_cell.font = Font(bold=True, color="856404")

    for column in ws.columns:
        max_len = max(len(str(c.value or "")) for c in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_len + 3, 20)

    # ── Summary sheet ──
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Exam ID").font = HEADER_FONT
    ws2.cell(row=1, column=2, value=exam_id)
    ws2.cell(row=2, column=1, value="Total Students").font = HEADER_FONT
    ws2.cell(row=2, column=2, value=len(student_results))
    ws2.cell(row=3, column=1, value="Questions").font = HEADER_FONT
    ws2.cell(row=3, column=2, value=len(question_nums))

    if student_results:
        scores = [s.get("totalScore", 0) for s in student_results]
        max_pts = student_results[0].get("totalMaxPoints", 0)
        ws2.cell(row=5, column=1, value="Average Score").font = HEADER_FONT
        ws2.cell(row=5, column=2, value=round(sum(scores) / len(scores), 2))
        ws2.cell(row=6, column=1, value="Highest Score").font = HEADER_FONT
        ws2.cell(row=6, column=2, value=round(max(scores), 2))
        ws2.cell(row=7, column=1, value="Lowest Score").font = HEADER_FONT
        ws2.cell(row=7, column=2, value=round(min(scores), 2))
        ws2.cell(row=8, column=1, value="Max Points").font = HEADER_FONT
        ws2.cell(row=8, column=2, value=max_pts)

        reviews = sum(
            1 for s in student_results
            if any(
                q.get("needsReview") or q.get("confidence", 1) < config.HIGH_CONF_THRESHOLD
                for q in s.get("questions", {}).values()
            )
        )
        ws2.cell(row=10, column=1, value="Papers Needing Review").font = HEADER_FONT
        ws2.cell(row=10, column=2, value=reviews)

    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"{exam_id}_results.xlsx")
    wb.save(filepath)
    print(f"[Export] Saved to {filepath}")
    return filepath
